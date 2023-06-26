import os


from openpyxl import load_workbook

class LoaderXlsx(object):

    def __init__(self, configData):

        self.errors = []
        self.log =[]
        self.configData = configData
        self.allowedFormTypes = ['subform','text','dropdown','radio', 'tickbox','checkbox','yn','ny','ynu','ynuo','multipanel', 'textbox', 'textarea','date', 'header', 'subheader', 'paragraph', 'datetime', 'time']
        self.allowedDataTypes = ['string', 'bool', 'float', 'int', 'header', 'complex', 'date','dateTime']
        self.allowedExtensions = ['subform','subject', 'image', 'session']
        self.allowedDropDownTypes = ['int', 'string']
        self.toggle_panels = ['ynuo','yn','ny','ynu', 'multipanel']
        self.textFormTypes =  ['header', 'subheader', 'paragraph', 'text']

    def hasError(self):
        return len(self.errors)

    def errors(self):
        return self.errors


    def filesList(self,path):
        out = []
        for file in os.listdir(path):
            if file.endswith('.xlsx') and '~' not in file:
                out.append(file)

        return out
        
    def loadExcel(self,path,fileName):
    
        fullPath = ('{}/{}' .format(path,fileName))
        wb = load_workbook(fullPath)
        out = {'schema': '', 'order' :[] , 'forms':dict()}

        formnames = []  # check for duplicates
        subformnames = []  # check for duplicates

        out['foundFormDetails'] = False
        
        for ws in wb:
            
            if ws.title == 'FormDetails':
                out['foundFormDetails'] = True
                schemaName = ws['B1'].value.replace(' ', '').replace("'","").replace('-', '')
                out['schema'] = schemaName.replace(' ', '').replace("'","").replace('-', '')


                out['restrict_to_project'] = ''
                if ws['A2'].value:
                    if 'Restrict to Project' in ws['A2'].value:
                        if ws['B2'].value:
                            restricted_proj = ws['B2'].value.replace(' ', '').replace("'", "").replace('-', '')
                            out['restrict_to_project'] = restricted_proj.replace(' ', '').replace("'", "").replace('-', '')


                for x in range(3,100):
                    
                    data = dict()
                   
                    strId = self.cleanCellValue(ws['A'+str(x)].value)
                    if 'Form' in strId:
                        formName = self.cleanCellValue(ws['B'+str(x)].value)
                        formRef = self.cleanCellValue(ws['C'+str(x)].value)
                        ext = self.cleanCellValue(ws['D'+str(x)].value)
                        desc = self.cleanCellValue(ws['E'+str(x)].value)
                        if 'ession' in '{}'.format(ext).lower():
                            ext = 'image'


                        if (strId) and (strId.startswith('Form')):

                            if not formName.strip() :
                                continue
                            else:

                                if not formName.strip() :
                                    self.errors.append('No Form ref, cell C is empty')
                                    continue
                                #replace spaces
                                strId = strId.replace(' ', '')
                                data['strId'] = strId
                                data['name'] = formName.replace("'","").replace('-', '')
                                data['ref'] = formRef.replace(' ', '').replace("'","").replace('-', '')
                                if not self.valid_name(data['ref']):
                                    self.errors.append(
                                        'Form {} is not valid - {} -  Check Form! '.format(
                                            strId, data['ref'] ))

                                data['inputs'] = []
                                data['ext'] = ext.lower()
                                data['desc'] = desc
                                data['schema'] = schemaName

                                out['order'].append(strId)
                                out['forms'][strId] = data

                                # check for duplicates

                                if data['ref'] in formnames and len(data['ref']) > 0:
                                    self.errors.append(
                                        'Duplicate!  Form {} is a duplicate: Check Form! '.format(
                                            strId ))
                                formnames.append(data['ref'])
                                if 'subform' in data['ext']:
                                    #subformnames
                                    subformnames.append(data['ref'])


        if not out['foundFormDetails']:
            self.errors.append('Page FormDetails page not found!!! Check CSV')

        if len(out['schema']) < 1:
            self.errors.append('Form {} Schema too short or missing: {} '.format(out['name'], out['schema']))


        for x in out['forms']:
            inputs = self.addInputs(x,wb,schemaName,out['forms'][x])

            if(len(inputs)) == 0:
                self.errors.append('Form {} has no inputs'.format(strId))
            out['forms'][x]['inputs'] = self.sortByDataName(inputs)

        #self.printForms(out['forms'])

        for x in out['forms']:
            self.checkForm(out['forms'][x],subformnames)

        return out

    def addInputs(self,strId,wb,schema,form):
        textnum = 1000
        out = []
        complex_type=False


        for ws in wb:
            
            if ws.title.replace(' ','') == strId:
                
                datatypes = [] # check for duplicates
                for x in range(2,500):
                    data =dict()

                    question = self.cleanCellValue(ws['A'+str(x)].value)
                    dataName = self.cleanCellValue(ws['B'+str(x)].value).replace(' ','')
                    subName = self.cleanCellValue(ws['C'+str(x)].value).replace(' ','')
                    formType = self.cleanCellValue(ws['D'+str(x)].value).replace(' ','')
                    dataType = self.cleanCellValue(ws['E'+str(x)].value).replace(' ','')
                    inputLen = self.setLength(ws['F'+str(x)].value)
                    opts = self.cleanDropDownOptions(self.cleanCellValue(ws['G'+str(x)].value))
                    hide = self.cleanCellValue(ws['H'+str(x)].value).replace(' ','')
                    required = self.cleanCellValue(ws['I' + str(x)].value).replace(' ', '')

                    if not question :
                        break

                    if not dataName:
                        if formType.lower() not in self.textFormTypes:
                            self.errors.append('Form {}, Input {} : has no Data Name'.format(strId,question))
                    if not self.valid_name(dataName):
                        self.errors.append('Form {}, Input {} : data name {} not valid - no special characters'.format(strId,question, dataName))

                    if dataName == "date":
                        # date not permiitted as dataname, creates erros
                        dataName = 'date{}'.format(schema)

                    if formType in self.textFormTypes:
                        subName = str(textnum)
                        textnum = textnum + 1
                    if 'heckbox' in formType:
                        formType = 'tickbox'

                    # need to loop for multiradio... REMOVED - just do checkbox list....
                    if 'multiradio' in formType:
                        mr_opts = opts.split(',')
                        mr_num=0
                        for mr_opt in mr_opts:
                            subName = '{}{}{}'.format(dataName, subName, mr_opt)
                            if mr_num < 1:
                                subName = '{}xstart'.format(subName)
                            if mr_num == len(mr_opts):
                                subName = '{}xend'.format(subName)
                            mr_num += 1



                            data['parent'] = ''
                            data['children'] = []
                            data['schema'] = schema
                            data['formRef'] = form['ref']

                            data['question'] = question
                            data['dataName'] = dataName
                            data['subName'] = subName
                            data['inputFormType']= formType.lower()
                            data['inputDataType']= dataType.lower()
                            data['inputLen']= self.setLength(inputLen)
                            data['required'] = "0"
                            if 'y' in required.lower() or 'x' in required.lower():
                                data['required'] = "1"


                            # check for duplicates
                            dt = '{}{}'.format(dataName, subName)
                            if dt in datatypes and len(dataName) > 0:
                                self.errors.append('Duplicate!  Form {}, Input {}  Dataname: {} Subname: {}: Check Form! '.format(strId,question, dataName, subName))
                            datatypes.append(dt)


                            data['hide']= hide
                            data['options']= opts

                            if data['inputFormType'] == 'dropdown' and not data['inputDataType']:
                                #only assumes int if specifies int
                                data['inputDataType'] = 'string'

                            # SHCEMA INPUT TYPE - if header, paragraph etc, schema type = text and is ignored
                            #print("{}{}".format(data['inputFormType'], data['inputDataType']))
                            data['schemaInputType'] = self.schemaInputType(data['inputFormType'], data['inputDataType'],data['options'])
                            if 'datetime' in formType:
                                #? any need for this?
                                data['inputDataType'] = 'dateTime'
                            #print('LOADER DATA: {} '.format(data))

                            self.checkInput(form['name'],data)
                            out.append(data)
                    else:
                        data['parent'] = ''
                        data['children'] = []
                        data['schema'] = schema
                        data['formRef'] = form['ref']

                        data['question'] = question
                        data['dataName'] = dataName
                        data['subName'] = subName
                        data['inputFormType'] = formType.lower()
                        data['inputDataType'] = dataType.lower()
                        data['inputLen'] = self.setLength(inputLen)
                        data['required'] = "0"
                        if 'y' in required.lower() or 'x' in required.lower():
                            data['required'] = "1"

                        # check for duplicates
                        dt = '{}{}'.format(dataName, subName)
                        if dt in datatypes and len(dataName) > 0:
                            self.errors.append(
                                'Duplicate!  Form {}, Input {}  Dataname: {} Subname: {}: Check Form! '.format(strId,
                                                                                                               question,
                                                                                                               dataName,
                                                                                                               subName))
                        datatypes.append(dt)

                        data['hide'] = hide
                        data['options'] = opts

                        if data['inputFormType'] == 'dropdown' and not data['inputDataType']:
                            # only assumes int if specifies int
                            data['inputDataType'] = 'string'

                        # SHCEMA INPUT TYPE - if header, paragraph etc, schema type = text and is ignored
                        # print("{}{}".format(data['inputFormType'], data['inputDataType']))
                        data['schemaInputType'] = self.schemaInputType(data['inputFormType'], data['inputDataType'],data['options'] )
                        if 'datetime' in formType:
                            # ? any need for this?
                            data['inputDataType'] = 'dateTime'
                        #print('LOADER DATA: {} '.format(data))

                        #self.checkInput(form['name'], data)
                        out.append(data)
                return out

    def checkInput(self,form,i, qs, subformnames=''):

        # check form type
        if i['inputFormType'] not in self.allowedFormTypes :
            self.errors.append('Form {} Question : {} ,Invalid Form Type {} '.format(form['name'], i['question'],i['inputFormType']))

        # check data type
        if (i['inputDataType'] != '') and (i['inputDataType'] not in self.allowedDataTypes) :
            self.errors.append('Form {} Question : {} ,Invalid Data Type {} '.format(form['name'], i['question'],i['inputDataType']))

        ## check to see if subforms exist, and options correct
        if 'subform' == i['inputFormType'] and i['options'] not in subformnames:
            self.errors.append('Form {} has subform with no exisiting type: {} - {} - {} '.format(form['name'], i['question'],i['options'], i['inputFormType']))

        # yn do not NEED subnames as already exist. (completed, reasonnotable)
        # ynuo has no subname
        if 'yn' == i['inputFormType'] and len(i['subName']) < 1:
            self.errors.append(
                'Form {} has missing subname: {} - {} - {}'.format(form['name'], i['question'], i['dataName'],
                                                                   i['inputFormType']))
        if 'ny' == i['inputFormType'] and len(i['subName']) < 1:
            self.errors.append(
                'Form {} has missing subname: {} - {} - {}'.format(form['name'], i['question'], i['dataName'],
                                                                   i['inputFormType']))
        if 'multipanel' == i['inputFormType']:
            opts = i['options'].split(',')
            if len(i['subName']) < 1:
                self.errors.append(
                    'Form {} has missing subname: {} - {} - {}'.format(form['name'], i['question'], i['dataName'],
                                                                       i['inputFormType']))
            if len(opts) != 2:
                self.errors.append(
                    'Form {} - multipanel question has wrong options format, see Wiki: {} - {} - {}'.format(
                        form['name'], i['question'], i['dataName'], i['options']))
            elif 'Y' not in opts[0] or 'N' not in opts[1]:
                self.errors.append(
                    'Form {} - multipanel question has wrong options format, see Wiki: {} - {} - {}'.format(
                        form['name'], i['question'], i['dataName'], i['options']))

        temp = i['question'].replace(" ", '') + i['dataName'] + i['subName']
        if temp in qs and i['inputFormType'] not in self.textFormTypes:
            self.errors.append('Form {} has duplicate input {} - {}'.format(form['name'], i['question'], i['dataName']))
        else:
            qs.append(temp)
        if len(i['children']) < 1 and i['inputFormType'] in self.toggle_panels:
            self.errors.append(
                'Form {}: Question: {} - Panel Toggle Type {} has no subtypes! '.format(form['name'], i['question'],
                                                                                        i['dataName']))
        # check form type
        if 'date' == i['inputFormType'] and i['inputLen'] != 0:
            self.errors.append('Form {} Question : {} , Length restriction applied to date type - not valid!'.format(form['name'], i['question'],i['inputFormType']))


    def schemaInputType(self,formType,dataType, options):
        #options should contain the subform name

        iType='none'
        if formType in self.textFormTypes:
            iType = 'text'
        else:
            iType = 'string' if not dataType else dataType

        #print('itype {}   form {}  data  {} formtype: {}'.format(iType, formType, dataType,self.textFormTypes))


        if formType in self.toggle_panels:
            iType = 'integer'
        elif formType == 'date':
            iType = 'date'
        elif formType == 'datetime':
            #iType = 'dateTime'
            # better as string??
            iType = 'string'


        elif formType == 'textbox' and 'int' not in dataType:
            if 'loat' not in dataType:
                #can be float as well. So change after
                iType = 'string'

        if 'ool' in dataType:
             iType='boolean'

        elif 'int' in dataType:
            iType='integer'
        elif 'loat' in dataType:
            iType='float'


        if formType == 'dropdown' and not dataType:
            #only assumes int if specifies int
            iType = 'string'

        if 'tickbox' in formType:
            iType='boolean'

        if formType == 'subform':
            ## need to check if existing form....
            iType=options

        return iType

    def setLength(self,data):
        if not data:
            return 0
        else:
            return int(data)

    def cleanCellValue(self,data):

        if (not data or data == None):
            
            return ''
        data = data.strip().encode('ascii', 'ignore')
        data_str=data.decode()

        return data_str

        # return '' if not data else data.strip()

    def valid_name(selfself, dataName):
        if set('[~!@#$%^&*()_+{}":;\']+$').intersection(dataName):
            return False
        else:
            return True


    def cleanDropDownOptions(self,o):

        a = o.split('-')

        if len(a) == 2 and ',' not in o and isinstance(a[0], int) and isinstance(a[1], int):
            r = a[0]
            for i in range(int(a[0]) + 1, int(a[1]) + 1):
                r = r + "," + str(i)

            print(r)
            return r
        
        bstr= '{}'.format(o)
        bstr = bstr.translate ({ord(c): "" for c in "!@#$%^&*()[]{};:./?\|`~=_"})
        
        #bstr = bstr.replace(" ","")

        return bstr

    def sortByDataName(self,inputs):

        temp = {'order': [], 'inputs' : {}}
        out = []
        #urgh, fix to stop recounting repeated paragraphs/headers
        textnum = 555
        for i in inputs:
            
            dataName = i['dataName']
            #### need to make so can be in panels
            if i['inputFormType'] in self.textFormTypes and len(i['dataName'])<1:
                dataName = str(textnum)
                textnum = textnum + 1

            if dataName in temp['inputs'].keys():
                i['parent'] = dataName
                temp['inputs'][dataName]['children'].append(i)
            else:
                temp['inputs'][dataName] = i 
                temp['order'].append(dataName)
                # out[i['dataName']] = [i]

        for o in temp['order']:
            out.append(temp['inputs'][o])
        #print('------------###########-----------')
        #print('{}'.format(out))
        #print temp['order']
        return out

    #TODO needs update
    def checkForm(self,form, subformnames):


        #print('CHECK FORMS: formanems:', subformnames)
        if form['ext'] not in self.allowedExtensions :
            self.errors.append('Form {} Invalid Extension {} '.format(form['name'], form['ext']))
        # already checked if no inputs
        # reset list for questions
        qs=[]

        for i in form['inputs']:
            self.checkInput( form, i,qs, subformnames)
            if len(i['children']) > 0:
                for c in i['children']:
                    self.checkInput(form, c,qs, subformnames)

            #TODO check for group inputs with no subname


    def printForms(self,forms):
       
        # already checked if no inputs
        for f in forms:
            print('Form {}'.format(forms[f]['name']))
            print('schema {}'.format(forms[f]['schema']))

            for i in forms[f]['inputs']:
                print(' -- input : question {}  : Type {}'.format(i['question'], i['inputFormType']))
                if len(i['children']) > 0:
                    for c in i['children']:
                        print('    -- input : question {} : Type {}'.format(c['question'],c['inputFormType']))


